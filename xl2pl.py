"""

Чтение и запись датафрейма polars.DataFrame в/из файла Excel

Автор: SaelKimberly: <sael.kimberly@yandex.ru>

"""
import re
import csv
from io import StringIO, BytesIO, BufferedReader
from pathlib import Path
from contextlib import closing
from typing import Union, Optional, Literal, Callable, Sequence, IO, cast

__dependencies__ = "polars numpy tqdm openpyxl".split()

for x in __dependencies__:
    try:
        print(f'{x:_<14}: {__import__(x, fromlist=["__version__"])}')
    except (ImportError, ModuleNotFoundError, AttributeError):
        raise RuntimeError(f"Для работы программы необходим модуль: {x}")


import polars as pl
import openpyxl as oxl

__all__ = ["load_excel", "save_excel"]


re_w = re.compile(r"\s+")


def load_excel(
    file: Union[str, Path, BufferedReader, BytesIO],
    sheet: Union[int, str] = 0,
    zero_cell: Optional[Union[Callable[[str], bool], Sequence[str], str]] = None,
    *,
    skip_rows: Optional[int] = None,
    skip_foot: Optional[int] = None,
    take_cols: Optional[Union[Callable[[str], bool], Sequence[str], str]] = None,
) -> pl.DataFrame:
    """
    Загружает файл Excel в датафрейм Polars

    :param file: Путь к файлу Excel, либо file-like объект формата Excel.
    :param sheet: Индекс или имя листа, подлежащего чтению в датафрейм.
    :param zero_cell: Заголовок первого столбца должен соответствовать условию:
    ((str) -> bool; "&lt;regexp&gt;"; либо [&lt;column-list&gt;])
    :param skip_rows: Пропустить &lt;skip_rows&gt; строк сверху
    :param skip_foot: Пропустить &lt;skip_foot&gt; строк снизу
    :param take_cols: Выбрать столбцы, соответствующие условию:
    ((str) -> bool; "&lt;regexp&gt;"; либо [&lt;column-list&gt;])

    :return: Датафрейм polars.DataFrame
    """

    # Проверка аргументов:
    if isinstance(file, str):
        path = Path(file)
    else:
        path = None

    if isinstance(file, (BufferedReader, BytesIO)):
        path = "<file-like>"  # type: ignore
    elif not isinstance(path, Path):
        raise ValueError(
            "Аргумент `file` должен быть строкой либо объектом "
            f"pathlib.Path, io.BufferedReader или io.BytesIO. Не {type(file)}."
        )
    elif path.suffix not in {"xlsx", "xls"}:
        raise ValueError(
            "Аргумент `path` должен содержать путь к файлу Excel "
            f"и заканчиваться расширением `xlsx` или `xls`, не `{path.suffix or ''}`."
        )
    elif not path.exists():
        raise ValueError(f"Файл не найден: {path}")

    if zero_cell is None:
        test_cell = lambda cell: True
    elif isinstance(zero_cell, str):
        test_cell = lambda cell: cell.value in {zero_cell}
    elif hasattr(zero_cell, "__iter__") and all(isinstance(x, str) for x in zero_cell):
        test_cell = lambda cell: cell.value in set(zero_cell)
    elif callable(zero_cell):
        test_cell = lambda cell: zero_cell(cell.value)  # type: ignore
    else:
        raise ValueError("test_cell")

    if take_cols is None:
        test_cols = lambda col: True
    elif isinstance(take_cols, str):
        test_cols = lambda col: bool(re.match(cast(str, take_cols), col))
    elif hasattr(take_cols, "__iter__") and all(isinstance(x, str) for x in take_cols):
        test_cols = lambda col: col in set(take_cols)
    elif callable(take_cols):
        test_cols = take_cols
    else:
        raise ValueError("take_cols")

    buffer = StringIO()
    writer = csv.writer(buffer, quoting=csv.QUOTE_ALL)

    read_immediately = zero_cell is None
    head_read = False
    i_row, i_col = 0, 0
    idxs = []

    try:
        book: oxl.Workbook = oxl.open(path, read_only=True, data_only=True)  # type: ignore
    except Exception as e:
        raise RuntimeError("Не удалось открыть файл Excel!") from e

    rows = []

    with closing(book):  # type: ignore
        try:
            if isinstance(sheet, int):
                _sheet = book.worksheets[0]
            elif isinstance(sheet, str):
                _sheet = book[sheet]
            else:
                raise ValueError("sheet")
        except ValueError:
            raise
        except Exception as e:
            raise ValueError(
                f"Не удалось загрузить лист `{sheet}` из файла Excel `{file}`!."
            ) from e

        for row in _sheet.iter_rows(min_row=(skip_rows or 0)):
            i_row += 1
            if not read_immediately:
                for i, cell in enumerate(row):
                    if test_cell(cell):
                        i_col = i
                        read_immediately = True
                        break

            if not read_immediately:
                continue

            if row[i_col].value is not None:
                if not head_read:
                    idxs = [
                        i_col + i for i, cell in enumerate(row) if test_cols(cell.value)
                    ]

                data = [row[i].value for i in idxs]

                if not head_read:
                    data = [re_w.sub(" ", old).strip() for old in data]
                    head_read = True

                rows.append(data)
            else:
                break

    if skip_foot is not None and len(rows) < 2 and skip_foot > (len(rows) - 1):
        raise ValueError(
            "Слишком большой параметр `skip_foot`: пропущены все строки датафрейма."
        )

    writer.writerows(rows if skip_foot is None else rows[:skip_foot])
    buffer.seek(0)

    return pl.read_csv(buffer, dtypes=[pl.Utf8 for _ in idxs])


def save_excel(
    path: Union[str, Path, IO[bytes]],
    df: pl.DataFrame,
    sheet_name: Optional[str] = None,
    *,
    if_sheet_exists: Literal["overwrite", "assert", "skip"] = "overwrite",
) -> pl.DataFrame:
    """
    Сохраняет датафрейм Polars в файл Excel (используя openpyxl)

    :param path: Путь к файлу, в котороый нужно сохранить датафрейм, либо file-like объект с возможностью записи
    :param df: Датафрейм polars.DataFrame
    :param sheet_name: Имя столбца (по-умолчанию: `Sheet`)
    :param if_sheet_exists: Если лист с именем <sheet_name> уже существует:
    `overwrite` - перезаписать (по-умолчанию);
    `assert` - выдать ошибку;
    `skip` - пропустить запись
    :return: Исходный датафрейм.
    """
    if isinstance(path, str):
        path = Path(path)
    if not isinstance(path, Path):
        raise ValueError(
            f"Аргумент `path` должен быть строкой или объектом pathlib.Path. Не {type(path)}."
        )
    if path.suffix == ".xlsx":
        raise ValueError(
            "Путь к файлу сохранения Excel должен заканчиваться "
            f"расширением `.xlsx`, не `{path.suffix or ''}`!"
        )

    if not path.exists():
        ctx = closing(oxl.workbook.Workbook(write_only=True))
    else:
        ctx = closing(oxl.open(path))  # type: ignore

    wb: oxl.Workbook
    with ctx as wb:
        sheets = wb.sheetnames[:]
        sheet_name = sheet_name or "Sheet"
        if sheet_name in sheets:
            if (
                not hasattr(if_sheet_exists, "__hash__")
                or if_sheet_exists.__hash__ is None
                or if_sheet_exists not in {"overwrite", "assert", "skip"}
            ):
                raise ValueError(
                    f"Аргумент `if_sheet_exists` не может быть `{if_sheet_exists}`. "
                    "Возможные варианты: `overwrite`, `assert` или `skip`."
                )
            if if_sheet_exists == "overwrite":
                del wb[sheet_name]
            elif if_sheet_exists == "assert":
                raise AssertionError(
                    f"Лист с именем `{sheet_name}` уже существует в файле `{path}`!"
                )
            else:
                # Не сохраняем в режиме 'skip', если лист уже существует.
                return df

        sheet = wb.create_sheet(sheet_name)

        head_wrot = False
        for row in df.to_dicts():
            if not head_wrot:
                sheet.append(list(iter(row.keys())))
                head_wrot = True
            sheet.append(list(iter(row.values())))
        try:
            wb.save(path)
        except Exception as e:
            raise RuntimeError(
                f"Не удалось сохранить файл {path}: ошибка доступа, или файл открыт в MS Excel!"
            ) from e

    return df
